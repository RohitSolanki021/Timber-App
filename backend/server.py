from fastapi import FastAPI, HTTPException, Depends, Header, Query, Body, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel, Field, EmailStr
from typing import Optional, List, Dict, Any, Literal
from datetime import datetime, timezone, timedelta
from pymongo import MongoClient, ASCENDING, DESCENDING
from bson import ObjectId
import os
import jwt
import hashlib
import secrets
import re
import json
import base64
import io

app = FastAPI(title="Natural Plylam Admin API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB connection
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "plylam_admin")
client = MongoClient(MONGO_URL)
db = client[DB_NAME]

JWT_SECRET = os.environ.get("JWT_SECRET", "plylam_secret_key_2024")
JWT_ALGORITHM = "HS256"

# Helper functions
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def create_token(user_id: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def verify_token(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid token")
    token = authorization.replace("Bearer ", "")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def serialize_doc(doc):
    if doc is None:
        return None
    doc["id"] = str(doc.pop("_id", ""))
    return doc

def get_next_customer_id():
    max_id = db.customers.find_one(sort=[("id", -1)])
    return (max_id["id"] + 1) if max_id and "id" in max_id else 1

def generate_order_id():
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    suffix = ''.join(secrets.choice(chars) for _ in range(8))
    return f"ORD-{suffix}"

def generate_invoice_id():
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    suffix = ''.join(secrets.choice(chars) for _ in range(8))
    return f"INV-{suffix}"

# Pydantic Models
class LoginRequest(BaseModel):
    email: str
    password: str
    app_role: Optional[str] = None

class RegisterRequest(BaseModel):
    name: str
    contactPerson: str
    phone: str
    email: str
    password: str

class CustomerCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    contactPerson: str = Field(..., min_length=1, max_length=100)
    email: str = Field(..., min_length=1)
    phone: str = Field(..., min_length=1, max_length=20)
    gst_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    pricing_tier: int = Field(default=1, ge=1, le=6)  # Changed to 6 tiers
    credit_limit: Optional[float] = 0
    notes: Optional[str] = None

class CustomerUpdate(BaseModel):
    name: Optional[str] = Field(None, min_length=1, max_length=200)
    contactPerson: Optional[str] = Field(None, min_length=1, max_length=100)
    email: Optional[str] = None
    phone: Optional[str] = Field(None, min_length=1, max_length=20)
    gst_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    pricing_tier: Optional[int] = Field(None, ge=1, le=6)
    credit_limit: Optional[float] = None
    notes: Optional[str] = None
    approval_status: Optional[str] = None
    is_active: Optional[bool] = None

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    phone: Optional[str] = None
    gst_number: Optional[str] = None
    approval_status: Optional[str] = None
    pricing_tier: Optional[int] = 1

class OrderStatusUpdate(BaseModel):
    order_id: str
    status: str

class InvoicePaidRequest(BaseModel):
    invoice_id: str

class ApproveCustomerRequest(BaseModel):
    customer_id: int

# MPIN login model
class MpinLoginRequest(BaseModel):
    phone: str
    mpin: str

class MpinSetRequest(BaseModel):
    mpin: str  # 4-6 digit PIN

# Banner model
class BannerCreate(BaseModel):
    title: str
    description: Optional[str] = None
    image_url: Optional[str] = None
    link_url: Optional[str] = None
    is_active: bool = True
    display_order: int = 0

class BannerUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    link_url: Optional[str] = None
    is_active: Optional[bool] = None
    display_order: Optional[int] = None

# Order Item model for the new direct order system
class OrderItemCreate(BaseModel):
    product_group: str  # "Plywood" or "Timber"
    product_id: str
    product_name: str
    thickness: str
    size: str
    quantity: int
    unit_price: float
    total_price: float

class DirectOrderCreate(BaseModel):
    customer_id: int
    items: List[OrderItemCreate]
    photo_reference: Optional[str] = None  # Base64 encoded image
    notes: Optional[str] = None
    # Transport details
    transport_mode: Optional[str] = None  # "Self Pickup" or "Delivery"
    vehicle_number: Optional[str] = None
    driver_name: Optional[str] = None
    driver_phone: Optional[str] = None

class OrderPriceUpdate(BaseModel):
    """Admin can update order prices before approval"""
    items: List[dict]  # Updated items with new prices
    notes: Optional[str] = None

class AdminCustomerCreate(BaseModel):
    """Admin/Super Admin creates a customer - auto approved"""
    name: str
    contactPerson: str
    email: str
    phone: str
    pricing_tier: int = Field(default=1, ge=1, le=6)
    gst_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    credit_limit: Optional[float] = 50000
    notes: Optional[str] = None

# Initialize demo data with new product catalog
def init_demo_data():
    # Clear existing data for fresh start
    if db.products_v2.count_documents({}) == 0:
        init_product_catalog()
    
    if db.stock.count_documents({}) == 0:
        init_stock_data()
    
    if db.users.count_documents({}) > 0:
        return
    
    admin = {
        "email": "admin@naturalplylam.com",
        "password": hash_password("admin123"),
        "name": "Admin User",
        "role": "Super Admin",
        "phone": "9876543210",
        "approval_status": "Approved",
        "pricing_tier": 1,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    db.users.insert_one(admin)
    
    manager = {
        "email": "manager@naturalplylam.com",
        "password": hash_password("manager123"),
        "name": "Manager User",
        "role": "Manager",
        "phone": "9876543211",
        "approval_status": "Approved",
        "pricing_tier": 1,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    db.users.insert_one(manager)
    
    worker_admin = {
        "email": "worker@naturalplylam.com",
        "password": hash_password("worker123"),
        "name": "Worker Admin",
        "role": "Admin",
        "phone": "9876543219",
        "approval_status": "Approved",
        "pricing_tier": 1,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    db.users.insert_one(worker_admin)
    
    sales_person = {
        "email": "sales@naturalplylam.com",
        "password": hash_password("sales123"),
        "name": "Rahul Sales",
        "role": "Sales Person",
        "phone": "9876543220",
        "approval_status": "Approved",
        "pricing_tier": 1,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    sales_result = db.users.insert_one(sales_person)
    sales_person_id = str(sales_result.inserted_id)
    
    # Customers with 6-tier pricing
    customers = [
        {"email": "customer1@example.com", "name": "ABC Furniture Works", "business_name": "ABC Furniture Works", "contactPerson": "John Doe", "phone": "9876543212", "role": "Customer", "approval_status": "Approved", "is_active": True, "pricing_tier": 2, "outstanding_balance": 15000, "credit_limit": 50000, "sales_person_id": sales_person_id, "sales_person_name": "Rahul Sales", "gst_number": "27AABCU9603R1ZM", "address": "123 Industrial Area, Mumbai, Maharashtra - 400001", "city": "Mumbai", "state": "Maharashtra", "pincode": "400001"},
        {"email": "customer2@example.com", "name": "XYZ Interiors", "business_name": "XYZ Interiors", "contactPerson": "Jane Smith", "phone": "9876543213", "role": "Customer", "approval_status": "Pending", "is_active": True, "pricing_tier": 1, "outstanding_balance": 0, "credit_limit": 25000, "sales_person_id": sales_person_id, "sales_person_name": "Rahul Sales", "address": "456 Commercial Complex, Delhi - 110001", "city": "Delhi", "state": "Delhi", "pincode": "110001"},
        {"email": "customer3@example.com", "name": "Modern Cabinets Ltd", "business_name": "Modern Cabinets Ltd", "contactPerson": "Mike Johnson", "phone": "9876543214", "role": "Customer", "approval_status": "Approved", "is_active": True, "pricing_tier": 3, "outstanding_balance": 25000, "credit_limit": 100000, "sales_person_id": sales_person_id, "sales_person_name": "Rahul Sales", "gst_number": "09AAACH7409R1ZZ", "address": "789 Manufacturing Hub, Bangalore - 560001", "city": "Bangalore", "state": "Karnataka", "pincode": "560001"},
        {"email": "customer4@example.com", "name": "Elite Woodworks", "business_name": "Elite Woodworks", "contactPerson": "Sarah Williams", "phone": "9876543215", "role": "Customer", "approval_status": "Approved", "is_active": True, "pricing_tier": 4, "outstanding_balance": 0, "credit_limit": 30000, "sales_person_id": sales_person_id, "sales_person_name": "Rahul Sales", "address": "321 Artisan Lane, Chennai - 600001", "city": "Chennai", "state": "Tamil Nadu", "pincode": "600001"},
        {"email": "customer5@example.com", "name": "Premium Plyboards", "business_name": "Premium Plyboards", "contactPerson": "David Brown", "phone": "9876543216", "role": "Customer", "approval_status": "Approved", "is_active": True, "pricing_tier": 5, "outstanding_balance": 8500, "credit_limit": 75000, "sales_person_id": sales_person_id, "sales_person_name": "Rahul Sales", "gst_number": "33AABCP1234A1ZX", "address": "567 Trade Center, Hyderabad - 500001", "city": "Hyderabad", "state": "Telangana", "pincode": "500001"},
        {"email": "customer6@example.com", "name": "Classic Interiors", "business_name": "Classic Interiors", "contactPerson": "Emily Davis", "phone": "9876543217", "role": "Customer", "approval_status": "Approved", "is_active": True, "pricing_tier": 6, "outstanding_balance": 0, "credit_limit": 40000, "sales_person_id": sales_person_id, "sales_person_name": "Rahul Sales", "address": "890 Design District, Pune - 411001", "city": "Pune", "state": "Maharashtra", "pincode": "411001"},
    ]
    for i, cust in enumerate(customers, 1):
        cust["id"] = i
        cust["password"] = hash_password("customer123")
        cust["created_at"] = datetime.now(timezone.utc).isoformat()
        cust["updated_at"] = datetime.now(timezone.utc).isoformat()
    db.customers.insert_many(customers)

def init_product_catalog():
    """Initialize the new product catalog with thickness and size variants"""
    
    # Product Groups
    product_groups = [
        {"id": "plywood", "name": "Plywood", "display_order": 1},
        {"id": "timber", "name": "Timber", "display_order": 2}
    ]
    db.product_groups.drop()
    db.product_groups.insert_many(product_groups)
    
    # Products with variants (from PDF catalog)
    products = [
        # Plywood Products
        {
            "id": "PP-BSL",
            "name": "PP BSL",
            "group": "Plywood",
            "description": "Premium Plus BSL Plywood",
            "base_price": 595.46,
            "price_unit": "sq.mt",
            "thicknesses": ["11"],
            "sizes": ["2.44 x 1.22"],
            "pricing_tiers": {
                "1": 595.46, "2": 580.00, "3": 565.00, 
                "4": 550.00, "5": 535.00, "6": 520.00
            }
        },
        {
            "id": "MDF-PP",
            "name": "MDF PP Plain",
            "group": "Plywood",
            "description": "MDF Premium Plus Plain",
            "base_price": 417.45,
            "price_unit": "sq.mt",
            "thicknesses": ["3", "5.5", "7.5", "11", "16", "16.75", "18", "25"],
            "sizes": ["2.44 x 1.22", "3.05 x 1.22"],
            "pricing_tiers": {
                "1": 417.45, "2": 400.00, "3": 385.00, 
                "4": 370.00, "5": 355.00, "6": 340.00
            }
        },
        # Timber Products  
        {
            "id": "HDF-BOIL",
            "name": "HDF Boil Plus Plain",
            "group": "Timber",
            "description": "HDF Boil Plus Plain Board",
            "base_price": 727.63,
            "price_unit": "sq.mt",
            "thicknesses": ["8", "12", "16.75", "18"],
            "sizes": ["2.44 x 1.22"],
            "pricing_tiers": {
                "1": 727.63, "2": 700.00, "3": 680.00, 
                "4": 660.00, "5": 640.00, "6": 620.00
            }
        },
        {
            "id": "MDF-DIR",
            "name": "MDF DIR Plain",
            "group": "Timber",
            "description": "MDF DIR Plain Board",
            "base_price": 260.61,
            "price_unit": "sq.mt",
            "thicknesses": ["1.9", "3.3", "5.5", "7", "7.3", "11", "14.5", "16.5", "18", "25"],
            "sizes": ["2.44 x 1.22"],
            "pricing_tiers": {
                "1": 260.61, "2": 250.00, "3": 240.00, 
                "4": 230.00, "5": 220.00, "6": 210.00
            }
        },
        {
            "id": "MDF-DWR",
            "name": "MDF DWR Plain",
            "group": "Timber",
            "description": "MDF DWR Plain Board",
            "base_price": 344.47,
            "price_unit": "sq.mt",
            "thicknesses": ["3.3", "5.5", "7.3", "11", "17", "18", "25"],
            "sizes": ["2.44 x 1.22"],
            "pricing_tiers": {
                "1": 344.47, "2": 330.00, "3": 315.00, 
                "4": 300.00, "5": 285.00, "6": 270.00
            }
        },
        {
            "id": "MDF-FSP",
            "name": "MDF FSP Plain",
            "group": "Timber",
            "description": "MDF FSP Plain Board",
            "base_price": 456.94,
            "price_unit": "sq.mt",
            "thicknesses": ["5.5", "8", "12", "16.75", "18"],
            "sizes": ["2.44 x 1.22"],
            "pricing_tiers": {
                "1": 456.94, "2": 440.00, "3": 425.00, 
                "4": 410.00, "5": 395.00, "6": 380.00
            }
        }
    ]
    db.products_v2.drop()
    db.products_v2.insert_many(products)

def init_stock_data():
    """Initialize stock data by product + thickness + size"""
    stock_entries = []
    
    products = list(db.products_v2.find({}))
    for product in products:
        for thickness in product.get("thicknesses", []):
            for size in product.get("sizes", []):
                stock_key = f"{product['id']}_{thickness}_{size.replace(' ', '').replace('x', 'X')}"
                stock_entries.append({
                    "stock_key": stock_key,
                    "product_id": product["id"],
                    "product_name": product["name"],
                    "group": product["group"],
                    "thickness": thickness,
                    "size": size,
                    "quantity": 100,  # Default stock
                    "reserved": 0,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })
    
    db.stock.drop()
    if stock_entries:
        db.stock.insert_many(stock_entries)

@app.on_event("startup")
async def startup_event():
    init_demo_data()

@app.get("/api/health")
async def health():
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}

# ============ AUTH ROUTES ============

@app.post("/api/login")
async def login(request: LoginRequest):
    # Customer login
    if request.app_role == "Customer":
        customer = db.customers.find_one({"email": request.email, "password": hash_password(request.password)})
        if not customer:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        if customer.get("approval_status") != "Approved":
            raise HTTPException(status_code=403, detail="Account not approved")
        if customer.get("is_active") == False:
            raise HTTPException(status_code=403, detail="Account is inactive")
        
        token = create_token(f"customer_{customer['id']}", "Customer")
        customer_data = {k: v for k, v in customer.items() if k not in ["_id", "password"]}
        return {"token": token, "user": customer_data}
    
    # Sales Person login
    if request.app_role == "Sales Person":
        user = db.users.find_one({"email": request.email, "password": hash_password(request.password), "role": "Sales Person"})
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        token = create_token(user["email"], user["role"])
        user_data = serialize_doc(user)
        del user_data["password"]
        return {"token": token, "user": user_data}
    
    # Admin/Manager login
    user = db.users.find_one({"email": request.email, "password": hash_password(request.password)})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["email"], user["role"])
    user_data = serialize_doc(user)
    del user_data["password"]
    
    return {"token": token, "user": user_data}

@app.post("/api/logout")
async def logout():
    return {"message": "Logged out successfully"}

@app.get("/api/me")
async def get_me(payload: dict = Depends(verify_token)):
    user_id = payload.get("user_id")
    
    if user_id.startswith("customer_"):
        customer_id = int(user_id.replace("customer_", ""))
        customer = db.customers.find_one({"id": customer_id}, {"_id": 0, "password": 0})
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        return customer
    
    # Admin/Manager - first try by email (new format), then by ObjectId (legacy)
    user = db.users.find_one({"email": user_id})
    if not user:
        try:
            user = db.users.find_one({"_id": ObjectId(user_id)})
        except:
            pass
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user_data = serialize_doc(user)
    del user_data["password"]
    return user_data

@app.post("/api/register")
async def register(request: RegisterRequest):
    existing = db.customers.find_one({"email": request.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    new_id = get_next_customer_id()
    new_customer = {
        "id": new_id,
        "email": request.email,
        "password": hash_password(request.password),
        "name": request.name,
        "business_name": request.name,
        "contactPerson": request.contactPerson,
        "phone": request.phone,
        "role": "Customer",
        "approval_status": "Pending",
        "is_active": True,
        "pricing_tier": 1,
        "outstanding_balance": 0,
        "credit_limit": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    db.customers.insert_one(new_customer)
    
    return {"success": True, "message": "Registration successful. Please wait for admin approval.", "customer_id": new_id}

# ============ MPIN LOGIN ============

@app.post("/api/mpin/set")
async def set_mpin(request: MpinSetRequest, payload: dict = Depends(verify_token)):
    """Set or update MPIN for quick login"""
    user_id = payload.get("user_id")
    mpin = request.mpin
    
    # Validate MPIN (4-6 digits)
    if not mpin.isdigit() or len(mpin) < 4 or len(mpin) > 6:
        raise HTTPException(status_code=400, detail="MPIN must be 4-6 digits")
    
    hashed_mpin = hash_password(mpin)
    
    if user_id.startswith("customer_"):
        customer_id = int(user_id.replace("customer_", ""))
        db.customers.update_one(
            {"id": customer_id},
            {"$set": {"mpin": hashed_mpin, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        db.users.update_one(
            {"email": user_id},
            {"$set": {"mpin": hashed_mpin, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    return {"success": True, "message": "MPIN set successfully"}

@app.post("/api/mpin/login")
async def mpin_login(request: MpinLoginRequest):
    """Login using phone number and MPIN"""
    phone = request.phone
    mpin_hash = hash_password(request.mpin)
    
    # Try customer login first
    customer = db.customers.find_one({"phone": phone, "mpin": mpin_hash})
    if customer:
        if customer.get("approval_status") != "Approved":
            raise HTTPException(status_code=403, detail="Account not approved")
        if customer.get("is_active") == False:
            raise HTTPException(status_code=403, detail="Account is inactive")
        
        token = create_token(f"customer_{customer['id']}", "Customer")
        customer_data = {k: v for k, v in customer.items() if k not in ["_id", "password", "mpin"]}
        return {"token": token, "user": customer_data}
    
    # Try user (admin/sales) login
    user = db.users.find_one({"phone": phone, "mpin": mpin_hash})
    if user:
        token = create_token(user["email"], user["role"])
        user_data = serialize_doc(user)
        del user_data["password"]
        if "mpin" in user_data:
            del user_data["mpin"]
        return {"token": token, "user": user_data}
    
    raise HTTPException(status_code=401, detail="Invalid phone number or MPIN")

@app.get("/api/mpin/check")
async def check_mpin_set(payload: dict = Depends(verify_token)):
    """Check if current user has MPIN set"""
    user_id = payload.get("user_id")
    
    if user_id.startswith("customer_"):
        customer_id = int(user_id.replace("customer_", ""))
        customer = db.customers.find_one({"id": customer_id})
        return {"has_mpin": bool(customer and customer.get("mpin"))}
    else:
        user = db.users.find_one({"email": user_id})
        return {"has_mpin": bool(user and user.get("mpin"))}

# ============ BANNER MANAGEMENT ============

@app.get("/api/banners")
async def get_banners():
    """Get all active banners (public endpoint)"""
    banners = list(db.banners.find({"is_active": True}, {"_id": 0}).sort("display_order", 1))
    return {"banners": banners}

@app.get("/api/admin/banners")
async def get_all_banners(payload: dict = Depends(verify_token)):
    """Get all banners (admin only)"""
    role = payload.get("role", "")
    if role not in ["Super Admin", "Admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    banners = list(db.banners.find({}, {"_id": 0}).sort("display_order", 1))
    return {"banners": banners}

@app.post("/api/admin/banners")
async def create_banner(banner: BannerCreate, payload: dict = Depends(verify_token)):
    """Create a new banner (admin only)"""
    role = payload.get("role", "")
    if role not in ["Super Admin", "Admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    banner_id = f"BNR-{secrets.token_hex(4).upper()}"
    banner_data = {
        "id": banner_id,
        "title": banner.title,
        "description": banner.description,
        "image_url": banner.image_url,
        "link_url": banner.link_url,
        "is_active": banner.is_active,
        "display_order": banner.display_order,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    db.banners.insert_one(banner_data)
    
    return {"success": True, "banner": {k: v for k, v in banner_data.items() if k != "_id"}}

@app.put("/api/admin/banners/{banner_id}")
async def update_banner(banner_id: str, banner: BannerUpdate, payload: dict = Depends(verify_token)):
    """Update a banner (admin only)"""
    role = payload.get("role", "")
    if role not in ["Super Admin", "Admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    existing = db.banners.find_one({"id": banner_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Banner not found")
    
    update_data = {k: v for k, v in banner.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    db.banners.update_one({"id": banner_id}, {"$set": update_data})
    
    updated = db.banners.find_one({"id": banner_id}, {"_id": 0})
    return {"success": True, "banner": updated}

@app.delete("/api/admin/banners/{banner_id}")
async def delete_banner(banner_id: str, payload: dict = Depends(verify_token)):
    """Delete a banner (admin only)"""
    role = payload.get("role", "")
    if role not in ["Super Admin", "Admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = db.banners.delete_one({"id": banner_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Banner not found")
    
    return {"success": True, "message": "Banner deleted"}

# ============ PRODUCT CATALOG ROUTES (New V2) ============

@app.get("/api/product-groups")
async def get_product_groups(payload: dict = Depends(verify_token)):
    """Get all product groups (Plywood, Timber)"""
    groups = list(db.product_groups.find({}, {"_id": 0}).sort("display_order", 1))
    return {"groups": groups}

@app.get("/api/products-v2")
async def get_products_v2(
    group: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    """Get products with their variants for the new order system"""
    query = {}
    if group:
        query["group"] = group
    
    products = list(db.products_v2.find(query, {"_id": 0}))
    return {"products": products}

@app.get("/api/products-v2/{product_id}")
async def get_product_v2(product_id: str, payload: dict = Depends(verify_token)):
    """Get a specific product with all its details"""
    product = db.products_v2.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"product": product}

@app.get("/api/products-v2/{product_id}/stock")
async def get_product_stock(
    product_id: str,
    thickness: Optional[str] = None,
    size: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    """Get stock for a product with optional thickness and size filters"""
    query = {"product_id": product_id}
    if thickness:
        query["thickness"] = thickness
    if size:
        query["size"] = size
    
    stock_items = list(db.stock.find(query, {"_id": 0}))
    return {"stock": stock_items}

@app.get("/api/stock/check")
async def check_stock(
    product_id: str,
    thickness: str,
    size: str,
    quantity: int,
    payload: dict = Depends(verify_token)
):
    """Check if requested quantity is available"""
    stock_key = f"{product_id}_{thickness}_{size.replace(' ', '').replace('x', 'X')}"
    stock = db.stock.find_one({"stock_key": stock_key})
    
    if not stock:
        return {"available": False, "message": "Product variant not found", "current_stock": 0}
    
    available_qty = stock.get("quantity", 0) - stock.get("reserved", 0)
    is_available = available_qty >= quantity
    
    return {
        "available": is_available,
        "current_stock": available_qty,
        "requested": quantity,
        "message": "In Stock" if is_available else f"Only {available_qty} available"
    }

@app.get("/api/price/calculate")
async def calculate_price(
    product_id: str,
    customer_id: int,
    quantity: int,
    payload: dict = Depends(verify_token)
):
    """Calculate price based on customer's pricing tier"""
    product = db.products_v2.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    customer = db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    pricing_tier = str(customer.get("pricing_tier", 1))
    unit_price = product.get("pricing_tiers", {}).get(pricing_tier, product.get("base_price", 0))
    total_price = unit_price * quantity
    
    return {
        "product_id": product_id,
        "product_name": product.get("name"),
        "pricing_tier": pricing_tier,
        "unit_price": unit_price,
        "quantity": quantity,
        "total_price": total_price,
        "price_unit": product.get("price_unit", "sq.mt")
    }

# ============ PRODUCT MANAGEMENT WITH VARIANTS ============

class ProductVariant(BaseModel):
    thickness: str
    size: str
    stock: int = 0
    prices: Dict[str, float]  # tier_num -> price

class ProductWithVariants(BaseModel):
    name: str
    group: str  # Plywood or Timber
    description: Optional[str] = ""
    variants: List[ProductVariant]

@app.post("/api/admin/products-v2")
async def create_product_v2(product: ProductWithVariants, payload: dict = Depends(verify_token)):
    """Create a new product with variants (thickness/size) and tier pricing"""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin"]:
        raise HTTPException(status_code=403, detail="Only admin can create products")
    
    # Generate product ID
    product_id = f"{product.group[:3].upper()}-{secrets.token_hex(4).upper()}"
    
    # Collect all unique thicknesses, sizes, and build pricing tiers
    thicknesses = list(set(v.thickness for v in product.variants))
    sizes = list(set(v.size for v in product.variants))
    
    # Use first variant's tier 1 price as base
    base_price = product.variants[0].prices.get("1", 0) if product.variants else 0
    pricing_tiers = product.variants[0].prices if product.variants else {}
    
    # Create the product
    product_doc = {
        "id": product_id,
        "name": product.name,
        "group": product.group,
        "description": product.description or "",
        "base_price": base_price,
        "price_unit": "piece",
        "thicknesses": sorted(thicknesses, key=lambda x: float(x) if x.replace('.','').isdigit() else 0),
        "sizes": sorted(sizes),
        "pricing_tiers": pricing_tiers,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    db.products_v2.insert_one(product_doc)
    
    # Create stock entries for each variant
    for variant in product.variants:
        stock_key = f"{product_id}_{variant.thickness}_{variant.size.replace(' ', '').replace('x', 'X')}"
        stock_doc = {
            "stock_key": stock_key,
            "product_id": product_id,
            "thickness": variant.thickness,
            "size": variant.size,
            "quantity": variant.stock,
            "reserved": 0,
            "variant_prices": variant.prices,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        db.stock.update_one(
            {"stock_key": stock_key},
            {"$set": stock_doc},
            upsert=True
        )
    
    del product_doc["_id"]
    return {"success": True, "message": "Product created", "product": product_doc}

@app.get("/api/admin/products-v2/template")
async def download_product_template(payload: dict = Depends(verify_token)):
    """Download Excel template for product import"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Headers
        headers = ["Product Name", "Group (Plywood/Timber)", "Description", "Thickness (mm)", "Size", "Stock Qty", 
                   "Tier 1 Price", "Tier 2 Price", "Tier 3 Price", "Tier 4 Price", "Tier 5 Price", "Tier 6 Price"]
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Sample data rows
        sample_data = [
            ["MR Plywood", "Plywood", "Moisture Resistant", "6", "8x4", 100, 500, 480, 460, 440, 420, 400],
            ["MR Plywood", "Plywood", "Moisture Resistant", "12", "8x4", 50, 800, 770, 740, 710, 680, 650],
            ["BWR Plywood", "Plywood", "Boiling Water Resistant", "19", "8x4", 75, 1200, 1150, 1100, 1050, 1000, 950],
            ["Teak Wood", "Timber", "Premium Teak", "25", "8x4", 30, 2500, 2400, 2300, 2200, 2100, 2000],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        column_widths = [20, 20, 25, 15, 10, 10, 12, 12, 12, 12, 12, 12]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width
        
        # Instructions sheet
        ws2 = wb.create_sheet("Instructions")
        instructions = [
            "PRODUCT IMPORT INSTRUCTIONS",
            "",
            "1. Product Name: Name of the product (e.g., 'MR Plywood', 'BWR Plywood')",
            "2. Group: Must be exactly 'Plywood' or 'Timber'",
            "3. Description: Optional product description",
            "4. Thickness: Thickness in mm (e.g., 6, 12, 19, 25)",
            "5. Size: Size format (e.g., '8x4', '10x4')",
            "6. Stock Qty: Initial stock quantity",
            "7-12. Tier Prices: Price for each customer tier (1-6)",
            "",
            "NOTES:",
            "- Each row represents one product variant (thickness + size combination)",
            "- Same product name can have multiple rows for different variants",
            "- All price fields are required",
            "- Do not change the header row",
        ]
        for idx, line in enumerate(instructions, 1):
            ws2.cell(row=idx, column=1, value=line)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=product_import_template.xlsx"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel libraries not installed")

@app.post("/api/admin/products-v2/import")
async def import_products_from_excel(file: UploadFile = File(...), payload: dict = Depends(verify_token)):
    """Import products from Excel file"""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin"]:
        raise HTTPException(status_code=403, detail="Only admin can import products")
    
    try:
        import pandas as pd
        
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
        
        # Expected columns
        required_cols = ["Product Name", "Group (Plywood/Timber)", "Thickness (mm)", "Size", "Stock Qty", 
                         "Tier 1 Price", "Tier 2 Price", "Tier 3 Price", "Tier 4 Price", "Tier 5 Price", "Tier 6 Price"]
        
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(status_code=400, detail=f"Missing columns: {missing_cols}")
        
        # Group rows by product name
        products_created = 0
        variants_created = 0
        errors = []
        
        grouped = df.groupby(["Product Name", "Group (Plywood/Timber)"])
        
        for (name, group), rows in grouped:
            # Generate product ID
            group_prefix = str(group)[:3].upper() if isinstance(group, str) else "PRD"
            product_id = f"{group_prefix}-{secrets.token_hex(4).upper()}"
            
            variants = []
            thicknesses = []
            sizes = []
            
            for _, row in rows.iterrows():
                try:
                    thickness = str(row["Thickness (mm)"])
                    size = str(row["Size"])
                    stock = int(row.get("Stock Qty", 0) or 0)
                    
                    prices = {}
                    for tier in range(1, 7):
                        price_col = f"Tier {tier} Price"
                        prices[str(tier)] = float(row.get(price_col, 0) or 0)
                    
                    thicknesses.append(thickness)
                    sizes.append(size)
                    
                    # Create stock entry
                    stock_key = f"{product_id}_{thickness}_{size.replace(' ', '').replace('x', 'X')}"
                    stock_doc = {
                        "stock_key": stock_key,
                        "product_id": product_id,
                        "thickness": thickness,
                        "size": size,
                        "quantity": stock,
                        "reserved": 0,
                        "variant_prices": prices,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    db.stock.update_one({"stock_key": stock_key}, {"$set": stock_doc}, upsert=True)
                    variants_created += 1
                    
                except Exception as e:
                    errors.append(f"Row error for {name}: {str(e)}")
            
            # Create product document
            first_row = rows.iloc[0]
            base_price = float(first_row.get("Tier 1 Price", 0) or 0)
            pricing_tiers = {str(i): float(first_row.get(f"Tier {i} Price", 0) or 0) for i in range(1, 7)}
            
            product_doc = {
                "id": product_id,
                "name": str(name),
                "group": str(group),
                "description": str(first_row.get("Description", "") or ""),
                "base_price": base_price,
                "price_unit": "piece",
                "thicknesses": sorted(list(set(thicknesses)), key=lambda x: float(x) if x.replace('.','').isdigit() else 0),
                "sizes": sorted(list(set(sizes))),
                "pricing_tiers": pricing_tiers,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            db.products_v2.insert_one(product_doc)
            products_created += 1
        
        return {
            "success": True,
            "message": f"Import completed",
            "products_created": products_created,
            "variants_created": variants_created,
            "errors": errors
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel libraries not installed")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import error: {str(e)}")

@app.get("/api/admin/products-v2/export")
async def export_products_to_excel(payload: dict = Depends(verify_token)):
    """Export all products to Excel"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Fetch all products and their stock/variants
        products = list(db.products_v2.find({}, {"_id": 0}))
        
        rows = []
        for product in products:
            stocks = list(db.stock.find({"product_id": product["id"]}, {"_id": 0}))
            
            for stock in stocks:
                prices = stock.get("variant_prices", product.get("pricing_tiers", {}))
                row = {
                    "Product Name": product["name"],
                    "Group (Plywood/Timber)": product["group"],
                    "Description": product.get("description", ""),
                    "Thickness (mm)": stock.get("thickness", ""),
                    "Size": stock.get("size", ""),
                    "Stock Qty": stock.get("quantity", 0),
                    "Tier 1 Price": prices.get("1", 0),
                    "Tier 2 Price": prices.get("2", 0),
                    "Tier 3 Price": prices.get("3", 0),
                    "Tier 4 Price": prices.get("4", 0),
                    "Tier 5 Price": prices.get("5", 0),
                    "Tier 6 Price": prices.get("6", 0),
                }
                rows.append(row)
        
        df = pd.DataFrame(rows)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Products')
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=products_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel libraries not installed")

# ============ DIRECT ORDER ROUTES ============

@app.post("/api/orders/direct")
async def create_direct_order(order: DirectOrderCreate, payload: dict = Depends(verify_token)):
    """Create a direct order - splits into plywood and timber bills internally"""
    
    # Validate customer
    customer = db.customers.find_one({"id": order.customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    if customer.get("approval_status") != "Approved":
        raise HTTPException(status_code=403, detail="Customer account not approved")
    
    # Get customer pricing tier
    pricing_tier = str(customer.get("pricing_tier", 1))
    
    # Separate items by group
    plywood_items = []
    timber_items = []
    
    for item in order.items:
        # Verify stock availability
        stock_key = f"{item.product_id}_{item.thickness}_{item.size.replace(' ', '').replace('x', 'X')}"
        stock = db.stock.find_one({"stock_key": stock_key})
        
        if not stock:
            raise HTTPException(status_code=400, detail=f"Product variant not found: {item.product_name} {item.thickness}mm {item.size}")
        
        available_qty = stock.get("quantity", 0) - stock.get("reserved", 0)
        if available_qty < item.quantity:
            raise HTTPException(status_code=400, detail=f"Insufficient stock for {item.product_name} {item.thickness}mm {item.size}. Available: {available_qty}")
        
        # Get correct price based on customer tier
        product = db.products_v2.find_one({"id": item.product_id})
        if product:
            tier_price = product.get("pricing_tiers", {}).get(pricing_tier, product.get("base_price", 0))
            item_dict = item.dict()
            item_dict["unit_price"] = tier_price
            item_dict["total_price"] = tier_price * item.quantity
        else:
            item_dict = item.dict()
        
        if item.product_group.lower() == "plywood":
            plywood_items.append(item_dict)
        else:
            timber_items.append(item_dict)
    
    # Create orders
    orders_created = []
    invoices_created = []
    current_time = datetime.now(timezone.utc).isoformat()
    
    # Get placer info
    user_id = payload.get("user_id")
    placer_role = payload.get("role")
    placer_name = None
    
    if user_id.startswith("customer_"):
        placer_name = customer.get("contactPerson", customer.get("name"))
    else:
        user = db.users.find_one({"email": user_id})
        if user:
            placer_name = user.get("name")
    
    # Transport details
    transport_info = {
        "transport_mode": order.transport_mode,
        "vehicle_number": order.vehicle_number,
        "driver_name": order.driver_name,
        "driver_phone": order.driver_phone
    }
    
    # Helper function to create invoice for an order
    def create_invoice_for_order(order_data: dict) -> dict:
        invoice_id = f"INV-{secrets.token_hex(4).upper()}"
        invoice = {
            "id": invoice_id,
            "order_id": order_data["id"],
            "customer_id": order_data["customer_id"],
            "customerName": order_data["customerName"],
            "order_type": order_data["order_type"],
            "items": order_data["items"],
            "sub_total": order_data["sub_total"],
            "cgst": order_data["cgst"],
            "sgst": order_data["sgst"],
            "grand_total": order_data["grand_total"],
            "pricing_tier": order_data["pricing_tier"],
            "status": "Pending",
            "issue_date": current_time,
            "due_date": (datetime.now(timezone.utc) + timedelta(days=30)).isoformat(),
            "created_at": current_time,
            "updated_at": current_time
        }
        db.invoices_v2.insert_one(invoice)
        return {"id": invoice_id, "type": order_data["order_type"], "total": order_data["grand_total"]}
    
    # Create Plywood order if items exist
    if plywood_items:
        plywood_total = sum(item["total_price"] for item in plywood_items)
        plywood_qty = sum(item["quantity"] for item in plywood_items)
        
        plywood_order = {
            "id": generate_order_id(),
            "customer_id": order.customer_id,
            "customerName": customer.get("name"),
            "order_type": "Plywood",
            "status": "Pending",  # Pending admin confirmation
            "items": plywood_items,
            "total_quantity": plywood_qty,
            "sub_total": plywood_total,
            "cgst": round(plywood_total * 0.09, 2),
            "sgst": round(plywood_total * 0.09, 2),
            "grand_total": round(plywood_total * 1.18, 2),
            "pricing_tier": pricing_tier,
            "photo_reference": order.photo_reference,
            "notes": order.notes,
            "placed_by": placer_name,
            "placed_by_role": placer_role,
            "transport": transport_info,
            "order_date": current_time,
            "created_at": current_time,
            "updated_at": current_time,
            "is_editable": True
        }
        db.orders_v2.insert_one(plywood_order)
        orders_created.append({"id": plywood_order["id"], "type": "Plywood", "total": plywood_order["grand_total"]})
        
        # Create invoice for plywood order
        inv = create_invoice_for_order(plywood_order)
        invoices_created.append(inv)
    
    # Create Timber order if items exist
    if timber_items:
        timber_total = sum(item["total_price"] for item in timber_items)
        timber_qty = sum(item["quantity"] for item in timber_items)
        
        timber_order = {
            "id": generate_order_id(),
            "customer_id": order.customer_id,
            "customerName": customer.get("name"),
            "order_type": "Timber",
            "status": "Pending",
            "items": timber_items,
            "total_quantity": timber_qty,
            "sub_total": timber_total,
            "cgst": round(timber_total * 0.09, 2),
            "sgst": round(timber_total * 0.09, 2),
            "grand_total": round(timber_total * 1.18, 2),
            "pricing_tier": pricing_tier,
            "photo_reference": order.photo_reference,
            "notes": order.notes,
            "placed_by": placer_name,
            "placed_by_role": placer_role,
            "transport": transport_info,
            "order_date": current_time,
            "created_at": current_time,
            "updated_at": current_time,
            "is_editable": True
        }
        db.orders_v2.insert_one(timber_order)
        orders_created.append({"id": timber_order["id"], "type": "Timber", "total": timber_order["grand_total"]})
        
        # Create invoice for timber order
        inv = create_invoice_for_order(timber_order)
        invoices_created.append(inv)
    
    total_amount = sum(o["total"] for o in orders_created)
    
    return {
        "success": True,
        "message": f"Order placed successfully. {len(orders_created)} order(s) and {len(invoices_created)} invoice(s) created.",
        "orders": orders_created,
        "invoices": invoices_created,
        "total_amount": total_amount
    }

@app.get("/api/orders/v2")
async def get_orders_v2(
    page: int = 1,
    per_page: int = 10,
    order_type: Optional[str] = None,
    status: Optional[str] = None,
    customer_id: Optional[int] = None,
    search: Optional[str] = None,
    pending_first: bool = True,
    payload: dict = Depends(verify_token)
):
    """Get orders with filters - pending orders shown first"""
    query = {}
    
    user_id = payload.get("user_id")
    role = payload.get("role")
    
    # Customer can only see their orders
    if user_id.startswith("customer_"):
        cust_id = int(user_id.replace("customer_", ""))
        query["customer_id"] = cust_id
    elif customer_id:
        query["customer_id"] = customer_id
    
    if order_type and order_type != "all":
        query["order_type"] = order_type
    
    if status and status != "all" and status != "All":
        query["status"] = status
    
    if search:
        query["$or"] = [
            {"id": {"$regex": search, "$options": "i"}},
            {"customerName": {"$regex": search, "$options": "i"}}
        ]
    
    total = db.orders_v2.count_documents(query)
    skip = (page - 1) * per_page
    
    # Sort: Pending first, then by created_at desc
    if pending_first and not status:
        pipeline = [
            {"$match": query},
            {"$addFields": {
                "sort_priority": {"$cond": [{"$eq": ["$status", "Pending"]}, 0, 1]}
            }},
            {"$sort": {"sort_priority": 1, "created_at": -1}},
            {"$skip": skip},
            {"$limit": per_page},
            {"$project": {"_id": 0, "sort_priority": 0}}
        ]
        orders = list(db.orders_v2.aggregate(pipeline))
    else:
        orders = list(db.orders_v2.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page))
    
    return {
        "data": orders,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }

@app.get("/api/orders/v2/{order_id}")
async def get_order_v2(order_id: str, payload: dict = Depends(verify_token)):
    """Get a specific order"""
    order = db.orders_v2.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"order": order}

# Customer-specific order endpoints
@app.get("/api/customer/orders")
async def get_customer_orders(
    page: int = 1,
    per_page: int = 10,
    order_type: Optional[str] = None,
    status: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    """Get orders for the logged-in customer"""
    user_id = payload.get("user_id")
    
    # Extract customer ID from token
    if user_id.startswith("customer_"):
        customer_id = int(user_id.replace("customer_", ""))
    else:
        # Try to find customer by email
        customer = db.customers.find_one({"email": user_id})
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        customer_id = customer["id"]
    
    query = {"customer_id": customer_id}
    
    if order_type and order_type != "all":
        query["order_type"] = order_type
    if status and status != "all" and status != "All":
        query["status"] = status
    
    total = db.orders_v2.count_documents(query)
    skip = (page - 1) * per_page
    
    orders = list(db.orders_v2.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page))
    
    return {
        "data": orders,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }

@app.get("/api/customer/orders/{order_id}")
async def get_customer_order(order_id: str, payload: dict = Depends(verify_token)):
    """Get a specific order for the logged-in customer"""
    user_id = payload.get("user_id")
    
    # Extract customer ID from token
    if user_id.startswith("customer_"):
        customer_id = int(user_id.replace("customer_", ""))
    else:
        customer = db.customers.find_one({"email": user_id})
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        customer_id = customer["id"]
    
    order = db.orders_v2.find_one({"id": order_id, "customer_id": customer_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return {"order": order, "data": order}

@app.put("/api/orders/v2/{order_id}")
async def update_order_v2(order_id: str, items: List[OrderItemCreate], payload: dict = Depends(verify_token)):
    """Update order items (only if order is editable)"""
    order = db.orders_v2.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if not order.get("is_editable", False):
        raise HTTPException(status_code=403, detail="Order is locked and cannot be edited")
    
    # Verify stock for new items
    for item in items:
        stock_key = f"{item.product_id}_{item.thickness}_{item.size.replace(' ', '').replace('x', 'X')}"
        stock = db.stock.find_one({"stock_key": stock_key})
        
        if not stock:
            raise HTTPException(status_code=400, detail=f"Product variant not found: {item.product_name}")
        
        available_qty = stock.get("quantity", 0) - stock.get("reserved", 0)
        if available_qty < item.quantity:
            raise HTTPException(status_code=400, detail=f"Insufficient stock for {item.product_name}. Available: {available_qty}")
    
    # Get customer pricing tier
    customer = db.customers.find_one({"id": order["customer_id"]})
    pricing_tier = str(customer.get("pricing_tier", 1)) if customer else "1"
    
    # Recalculate with correct prices
    updated_items = []
    for item in items:
        product = db.products_v2.find_one({"id": item.product_id})
        if product:
            tier_price = product.get("pricing_tiers", {}).get(pricing_tier, product.get("base_price", 0))
            item_dict = item.dict()
            item_dict["unit_price"] = tier_price
            item_dict["total_price"] = tier_price * item.quantity
        else:
            item_dict = item.dict()
        updated_items.append(item_dict)
    
    total = sum(item["total_price"] for item in updated_items)
    total_qty = sum(item["quantity"] for item in updated_items)
    
    db.orders_v2.update_one(
        {"id": order_id},
        {"$set": {
            "items": updated_items,
            "total_quantity": total_qty,
            "sub_total": total,
            "cgst": round(total * 0.09, 2),
            "sgst": round(total * 0.09, 2),
            "grand_total": round(total * 1.18, 2),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "message": "Order updated successfully"}

class AdminOrderItemsUpdate(BaseModel):
    items: List[dict]
    notes: Optional[str] = None

@app.put("/api/orders/v2/{order_id}/items")
async def update_order_items_v2(order_id: str, body: AdminOrderItemsUpdate, payload: dict = Depends(verify_token)):
    """Admin can update order items and prices before approval"""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager", "sales person"]:
        raise HTTPException(status_code=403, detail="Not authorized to update order prices")
    
    order = db.orders_v2.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.get("status") != "Pending":
        raise HTTPException(status_code=400, detail="Can only update pending orders")
    
    items = body.items
    # Update items with admin-provided prices
    total = sum(item.get("total_price", item.get("unit_price", 0) * item.get("quantity", 1)) for item in items)
    total_qty = sum(item.get("quantity", 1) for item in items)
    
    update_data = {
        "items": items,
        "total_quantity": total_qty,
        "sub_total": total,
        "cgst": round(total * 0.09, 2),
        "sgst": round(total * 0.09, 2),
        "grand_total": round(total * 1.18, 2),
        "price_modified_by": payload.get("user_id"),
        "price_modified_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if body.notes:
        update_data["admin_notes"] = body.notes
    
    db.orders_v2.update_one({"id": order_id}, {"$set": update_data})
    
    return {"success": True, "message": "Order prices updated", "new_total": round(total * 1.18, 2)}

@app.put("/api/admin/orders/{order_id}/items")
async def admin_update_order_items(order_id: str, body: AdminOrderItemsUpdate, payload: dict = Depends(verify_token)):
    """Admin can update order items and prices before approval"""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admin can update order prices")
    
    order = db.orders_v2.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.get("status") != "Pending":
        raise HTTPException(status_code=400, detail="Can only update pending orders")
    
    items = body.items
    # Update items with admin-provided prices
    total = sum(item.get("total_price", item.get("unit_price", 0) * item.get("quantity", 1)) for item in items)
    total_qty = sum(item.get("quantity", 1) for item in items)
    
    update_data = {
        "items": items,
        "total_quantity": total_qty,
        "sub_total": total,
        "cgst": round(total * 0.09, 2),
        "sgst": round(total * 0.09, 2),
        "grand_total": round(total * 1.18, 2),
        "price_modified_by": payload.get("user_id"),
        "price_modified_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if body.notes:
        update_data["admin_notes"] = body.notes
    
    db.orders_v2.update_one({"id": order_id}, {"$set": update_data})
    
    # Also update the corresponding invoice
    db.invoices_v2.update_one(
        {"order_id": order_id},
        {"$set": {
            "items": items,
            "sub_total": total,
            "cgst": round(total * 0.09, 2),
            "sgst": round(total * 0.09, 2),
            "grand_total": round(total * 1.18, 2),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "message": "Order prices updated", "new_total": round(total * 1.18, 2)}

@app.post("/api/orders/v2/{order_id}/confirm")
async def confirm_order_v2(order_id: str, payload: dict = Depends(verify_token)):
    """Admin confirms an order - locks it and deducts stock"""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admin can confirm orders")
    
    order = db.orders_v2.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.get("status") != "Pending":
        raise HTTPException(status_code=400, detail="Order is not in pending status")
    
    # Deduct stock
    for item in order.get("items", []):
        stock_key = f"{item['product_id']}_{item['thickness']}_{item['size'].replace(' ', '').replace('x', 'X')}"
        db.stock.update_one(
            {"stock_key": stock_key},
            {"$inc": {"quantity": -item["quantity"]}}
        )
    
    # Update order status
    db.orders_v2.update_one(
        {"id": order_id},
        {"$set": {
            "status": "Approved",
            "is_editable": False,
            "confirmed_at": datetime.now(timezone.utc).isoformat(),
            "confirmed_by": payload.get("user_id"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update invoice status (invoice created when order placed)
    db.invoices_v2.update_one(
        {"order_id": order_id},
        {"$set": {
            "status": "Approved",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    invoice = db.invoices_v2.find_one({"order_id": order_id}, {"_id": 0})
    invoice_id = invoice.get("id") if invoice else None
    
    return {"success": True, "message": "Order confirmed", "invoice_id": invoice_id}

@app.put("/api/orders/v2/{order_id}/price")
async def update_order_price(order_id: str, payload: dict = Depends(verify_token)):
    """Admin updates order prices before approval (from request body)"""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admin can update prices")
    
    order = db.orders_v2.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.get("status") != "Pending":
        raise HTTPException(status_code=400, detail="Can only update prices for pending orders")
    
    return {"success": True, "message": "Use PUT /api/orders/v2/{order_id}/items to update items with new prices"}

@app.post("/api/orders/v2/{order_id}/cancel")
async def cancel_order_v2(order_id: str, payload: dict = Depends(verify_token)):
    """Cancel an order"""
    order = db.orders_v2.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # If order was confirmed, restore stock
    if order.get("status") == "Approved":
        for item in order.get("items", []):
            stock_key = f"{item['product_id']}_{item['thickness']}_{item['size'].replace(' ', '').replace('x', 'X')}"
            db.stock.update_one(
                {"stock_key": stock_key},
                {"$inc": {"quantity": item["quantity"]}}
            )
    
    db.orders_v2.update_one(
        {"id": order_id},
        {"$set": {
            "status": "Cancelled",
            "is_editable": False,
            "cancelled_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "message": "Order cancelled"}

@app.put("/api/orders/v2/{order_id}/status")
async def update_order_status_v2(order_id: str, status: str = Body(..., embed=True), payload: dict = Depends(verify_token)):
    """Update order status (for Dispatched, Completed, etc.)"""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager", "sales person"]:
        raise HTTPException(status_code=403, detail="Only admin can update order status")
    
    order = db.orders_v2.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    valid_statuses = ["Pending", "Approved", "Delivered", "Cancelled"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    update_data = {
        "status": status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Mark as non-editable for certain statuses
    if status in ["Approved", "Delivered", "Cancelled"]:
        update_data["is_editable"] = False
    
    db.orders_v2.update_one({"id": order_id}, {"$set": update_data})
    
    return {"success": True, "message": f"Order status updated to {status}"}

# ============ INVOICE ROUTES (V2) ============

@app.get("/api/invoices/v2")
async def get_invoices_v2(
    page: int = 1,
    per_page: int = 10,
    order_type: Optional[str] = None,
    status: Optional[str] = None,
    customer_id: Optional[int] = None,
    payload: dict = Depends(verify_token)
):
    """Get invoices with filters"""
    query = {}
    
    user_id = payload.get("user_id")
    if user_id.startswith("customer_"):
        cust_id = int(user_id.replace("customer_", ""))
        query["customer_id"] = cust_id
    elif customer_id:
        query["customer_id"] = customer_id
    
    if order_type and order_type != "all":
        query["order_type"] = order_type
    
    if status and status != "all":
        query["status"] = status
    
    total = db.invoices_v2.count_documents(query)
    skip = (page - 1) * per_page
    
    invoices = list(db.invoices_v2.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page))
    
    return {
        "data": invoices,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }

@app.get("/api/invoices/v2/{invoice_id}")
async def get_invoice_v2(invoice_id: str, payload: dict = Depends(verify_token)):
    """Get a specific invoice"""
    invoice = db.invoices_v2.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"invoice": invoice}

# Customer-specific invoice endpoints
@app.get("/api/customer/invoices")
async def get_customer_invoices(
    page: int = 1,
    per_page: int = 10,
    order_type: Optional[str] = None,
    status: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    """Get invoices for the logged-in customer"""
    user_id = payload.get("user_id")
    
    # Extract customer ID from token
    if user_id.startswith("customer_"):
        customer_id = int(user_id.replace("customer_", ""))
    else:
        customer = db.customers.find_one({"email": user_id})
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        customer_id = customer["id"]
    
    query = {"customer_id": customer_id}
    
    if order_type and order_type != "all":
        query["order_type"] = order_type
    if status and status != "all" and status != "All":
        query["status"] = status
    
    total = db.invoices_v2.count_documents(query)
    skip = (page - 1) * per_page
    
    invoices = list(db.invoices_v2.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page))
    
    return {
        "data": invoices,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }

@app.get("/api/customer/invoices/{invoice_id}")
async def get_customer_invoice(invoice_id: str, payload: dict = Depends(verify_token)):
    """Get a specific invoice for the logged-in customer"""
    user_id = payload.get("user_id")
    
    # Extract customer ID from token
    if user_id.startswith("customer_"):
        customer_id = int(user_id.replace("customer_", ""))
    else:
        customer = db.customers.find_one({"email": user_id})
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        customer_id = customer["id"]
    
    invoice = db.invoices_v2.find_one({"id": invoice_id, "customer_id": customer_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"invoice": invoice, "data": invoice}

@app.put("/api/invoices/v2/{invoice_id}/status")
async def update_invoice_status_v2(invoice_id: str, status: str = Body(..., embed=True), payload: dict = Depends(verify_token)):
    """Update invoice status"""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admin can update invoice status")
    
    valid_statuses = ["Pending", "Paid", "Partially Paid", "Overdue", "Cancelled"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    result = db.invoices_v2.update_one(
        {"id": invoice_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"success": True, "message": f"Invoice status updated to {status}"}

# ============ PDF GENERATION ============

@app.get("/api/invoices/v2/{invoice_id}/pdf")
async def generate_invoice_pdf(invoice_id: str, payload: dict = Depends(verify_token)):
    """Generate PDF for an invoice (browser-based HTML that can be printed)"""
    invoice = db.invoices_v2.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    customer = db.customers.find_one({"id": invoice.get("customer_id")}, {"_id": 0, "password": 0})
    
    # Generate HTML for PDF
    items_html = ""
    for i, item in enumerate(invoice.get("items", []), 1):
        items_html += f"""
        <tr>
            <td>{i}</td>
            <td>{item.get('product_name', '')}</td>
            <td>{item.get('thickness', '')}mm</td>
            <td>{item.get('size', '')}</td>
            <td>{item.get('quantity', 0)}</td>
            <td>₹{item.get('unit_price', 0):,.2f}</td>
            <td>₹{item.get('total_price', 0):,.2f}</td>
        </tr>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Invoice {invoice['id']}</title>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; }}
            .header {{ display: flex; justify-content: space-between; margin-bottom: 30px; }}
            .company {{ font-size: 24px; font-weight: bold; color: #c00; }}
            .invoice-title {{ text-align: right; }}
            .invoice-title h1 {{ margin: 0; color: #333; }}
            .details {{ display: flex; justify-content: space-between; margin-bottom: 20px; }}
            .bill-to, .invoice-info {{ width: 48%; }}
            .bill-to h3, .invoice-info h3 {{ color: #666; margin-bottom: 10px; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
            th {{ background: #f5f5f5; }}
            .totals {{ text-align: right; margin-top: 20px; }}
            .totals table {{ width: 300px; margin-left: auto; }}
            .grand-total {{ font-size: 18px; font-weight: bold; background: #c00; color: white; }}
            .order-type {{ display: inline-block; padding: 4px 12px; border-radius: 4px; font-weight: bold; 
                           background: {('#ff9800' if invoice.get('order_type') == 'Plywood' else '#4caf50')}; color: white; }}
            @media print {{
                body {{ padding: 0; }}
                button {{ display: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company">Natural Plylam</div>
            <div class="invoice-title">
                <h1>INVOICE</h1>
                <p><span class="order-type">{invoice.get('order_type', 'Order')}</span></p>
            </div>
        </div>
        
        <div class="details">
            <div class="bill-to">
                <h3>Bill To:</h3>
                <p><strong>{customer.get('name', '') if customer else invoice.get('customerName', '')}</strong></p>
                <p>{customer.get('address', '') if customer else ''}</p>
                <p>GST: {customer.get('gst_number', 'N/A') if customer else 'N/A'}</p>
            </div>
            <div class="invoice-info">
                <h3>Invoice Details:</h3>
                <p><strong>Invoice #:</strong> {invoice['id']}</p>
                <p><strong>Order #:</strong> {invoice.get('order_id', '')}</p>
                <p><strong>Date:</strong> {invoice.get('issue_date', '')}</p>
                <p><strong>Due Date:</strong> {invoice.get('due_date', '')}</p>
                <p><strong>Status:</strong> {invoice.get('status', '')}</p>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th>Product</th>
                    <th>Thickness</th>
                    <th>Size</th>
                    <th>Qty</th>
                    <th>Unit Price</th>
                    <th>Total</th>
                </tr>
            </thead>
            <tbody>
                {items_html}
            </tbody>
        </table>
        
        <div class="totals">
            <table>
                <tr>
                    <td>Sub Total:</td>
                    <td>₹{invoice.get('sub_total', 0):,.2f}</td>
                </tr>
                <tr>
                    <td>CGST (9%):</td>
                    <td>₹{invoice.get('cgst', 0):,.2f}</td>
                </tr>
                <tr>
                    <td>SGST (9%):</td>
                    <td>₹{invoice.get('sgst', 0):,.2f}</td>
                </tr>
                <tr class="grand-total">
                    <td>Grand Total:</td>
                    <td>₹{invoice.get('grand_total', 0):,.2f}</td>
                </tr>
            </table>
        </div>
        
        <button onclick="window.print()" style="margin-top: 20px; padding: 10px 20px; background: #c00; color: white; border: none; cursor: pointer; border-radius: 4px;">
            Print / Download PDF
        </button>
    </body>
    </html>
    """
    
    return Response(content=html, media_type="text/html")

# ============ ADMIN DASHBOARD ============

@app.get("/api/admin/dashboard/v2")
async def get_admin_dashboard_v2(payload: dict = Depends(verify_token)):
    """Dashboard with separate plywood and timber sections"""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get counts
    pending_orders = db.orders_v2.count_documents({"status": "Pending"})
    confirmed_orders = db.orders_v2.count_documents({"status": "Approved"})
    
    plywood_orders = db.orders_v2.count_documents({"order_type": "Plywood"})
    timber_orders = db.orders_v2.count_documents({"order_type": "Timber"})
    
    pending_invoices = db.invoices_v2.count_documents({"status": "Pending"})
    
    total_customers = db.customers.count_documents({"approval_status": "Approved"})
    pending_customers = db.customers.count_documents({"approval_status": "Pending"})
    
    # Latest orders - pending first
    latest_plywood = list(db.orders_v2.aggregate([
        {"$match": {"order_type": "Plywood"}},
        {"$addFields": {"sort_priority": {"$cond": [{"$eq": ["$status", "Pending"]}, 0, 1]}}},
        {"$sort": {"sort_priority": 1, "created_at": -1}},
        {"$limit": 10},
        {"$project": {"_id": 0, "sort_priority": 0}}
    ]))
    
    latest_timber = list(db.orders_v2.aggregate([
        {"$match": {"order_type": "Timber"}},
        {"$addFields": {"sort_priority": {"$cond": [{"$eq": ["$status", "Pending"]}, 0, 1]}}},
        {"$sort": {"sort_priority": 1, "created_at": -1}},
        {"$limit": 10},
        {"$project": {"_id": 0, "sort_priority": 0}}
    ]))
    
    return {
        "pending_orders": pending_orders,
        "confirmed_orders": confirmed_orders,
        "plywood_orders_count": plywood_orders,
        "timber_orders_count": timber_orders,
        "pending_invoices": pending_invoices,
        "total_customers": total_customers,
        "pending_customers": pending_customers,
        "latest_plywood_orders": latest_plywood,
        "latest_timber_orders": latest_timber
    }

@app.get("/api/admin/analytics/products")
async def get_product_analytics(
    customer_id: Optional[int] = None,
    product_id: Optional[str] = None,
    order_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    """Analytics: Product sales by customer, quantity, etc."""
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    match_stage = {}
    if customer_id:
        match_stage["customer_id"] = customer_id
    if order_type:
        match_stage["order_type"] = order_type
    if start_date:
        match_stage["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in match_stage:
            match_stage["created_at"]["$lte"] = end_date
        else:
            match_stage["created_at"] = {"$lte": end_date}
    
    # Aggregate sales by product
    pipeline = [
        {"$match": match_stage} if match_stage else {"$match": {}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": {
                "product_id": "$items.product_id",
                "product_name": "$items.product_name",
                "product_group": "$items.product_group"
            },
            "total_quantity": {"$sum": "$items.quantity"},
            "total_amount": {"$sum": "$items.total_price"},
            "order_count": {"$sum": 1}
        }},
        {"$sort": {"total_quantity": -1}},
        {"$project": {
            "_id": 0,
            "product_id": "$_id.product_id",
            "product_name": "$_id.product_name",
            "product_group": "$_id.product_group",
            "total_quantity": 1,
            "total_amount": 1,
            "order_count": 1
        }}
    ]
    
    if product_id:
        pipeline[0] = {"$match": {**match_stage, "items.product_id": product_id}}
    
    product_stats = list(db.orders_v2.aggregate(pipeline))
    
    # Sales by customer
    customer_pipeline = [
        {"$match": match_stage} if match_stage else {"$match": {}},
        {"$group": {
            "_id": {"customer_id": "$customer_id", "customerName": "$customerName"},
            "total_orders": {"$sum": 1},
            "total_quantity": {"$sum": "$total_quantity"},
            "total_amount": {"$sum": "$grand_total"}
        }},
        {"$sort": {"total_amount": -1}},
        {"$limit": 20},
        {"$project": {
            "_id": 0,
            "customer_id": "$_id.customer_id",
            "customer_name": "$_id.customerName",
            "total_orders": 1,
            "total_quantity": 1,
            "total_amount": 1
        }}
    ]
    
    customer_stats = list(db.orders_v2.aggregate(customer_pipeline))
    
    return {
        "product_sales": product_stats,
        "customer_sales": customer_stats
    }

# ============ KEEP EXISTING ROUTES FOR BACKWARD COMPATIBILITY ============
# (Customer routes, old product routes, old order routes - keeping them for the admin panel)

@app.get("/api/customers")
async def get_customers(
    page: int = 1,
    per_page: int = 10,
    search: Optional[str] = None,
    status: Optional[str] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    pending_first: bool = True,
    payload: dict = Depends(verify_token)
):
    query = {}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"contactPerson": {"$regex": search, "$options": "i"}},
            {"gst_number": {"$regex": search, "$options": "i"}}
        ]
    
    if status and status != "all":
        if status == "active":
            query["approval_status"] = "Approved"
            query["is_active"] = True
        elif status == "inactive":
            query["approval_status"] = "Approved"
            query["is_active"] = False
        elif status == "pending":
            query["approval_status"] = "Pending"
        elif status == "archived":
            query["approval_status"] = "Archived"
    
    total = db.customers.count_documents(query)
    skip = (page - 1) * per_page
    
    # Sort: Pending first, then by sort_by
    if pending_first and not status:
        # Use aggregation to sort pending first
        pipeline = [
            {"$match": query},
            {"$addFields": {
                "sort_priority": {"$cond": [{"$eq": ["$approval_status", "Pending"]}, 0, 1]}
            }},
            {"$sort": {"sort_priority": 1, sort_by: -1 if sort_order == "desc" else 1}},
            {"$skip": skip},
            {"$limit": per_page},
            {"$project": {"_id": 0, "password": 0, "mpin": 0, "sort_priority": 0}}
        ]
        customers = list(db.customers.aggregate(pipeline))
    else:
        sort_direction = DESCENDING if sort_order == "desc" else ASCENDING
        customers = list(db.customers.find(query, {"_id": 0, "password": 0, "mpin": 0}).sort(sort_by, sort_direction).skip(skip).limit(per_page))
    
    return {
        "data": customers,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }

@app.get("/api/customers/{customer_id}")
async def get_customer(customer_id: int, payload: dict = Depends(verify_token)):
    customer = db.customers.find_one({"id": customer_id}, {"_id": 0, "password": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"customer": customer}

@app.post("/api/customers")
async def create_customer(customer: CustomerCreate, payload: dict = Depends(verify_token)):
    existing = db.customers.find_one({"email": customer.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    new_id = get_next_customer_id()
    new_customer = {
        "id": new_id,
        **customer.dict(),
        "role": "Customer",
        "approval_status": "Approved",
        "is_active": True,
        "outstanding_balance": 0,
        "password": hash_password("customer123"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Handle pricing_tier vs pricing_type compatibility
    if "pricing_type" in new_customer:
        new_customer["pricing_tier"] = new_customer.pop("pricing_type", 1)
    
    db.customers.insert_one(new_customer)
    del new_customer["_id"]
    del new_customer["password"]
    
    return {"success": True, "customer": new_customer}

@app.put("/api/customers/{customer_id}")
async def update_customer(customer_id: int, customer: CustomerUpdate, payload: dict = Depends(verify_token)):
    existing = db.customers.find_one({"id": customer_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    update_data = {k: v for k, v in customer.dict().items() if v is not None}
    
    # Handle pricing_tier vs pricing_type compatibility
    if "pricing_type" in update_data:
        update_data["pricing_tier"] = update_data.pop("pricing_type")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    db.customers.update_one({"id": customer_id}, {"$set": update_data})
    
    updated = db.customers.find_one({"id": customer_id}, {"_id": 0, "password": 0})
    return {"success": True, "customer": updated}

@app.post("/api/customers/{customer_id}/approve")
async def approve_customer(customer_id: int, payload: dict = Depends(verify_token)):
    result = db.customers.update_one(
        {"id": customer_id},
        {"$set": {"approval_status": "Approved", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"success": True, "message": "Customer approved"}

@app.post("/api/customers/{customer_id}/reject")
async def reject_customer(customer_id: int, payload: dict = Depends(verify_token)):
    result = db.customers.update_one(
        {"id": customer_id},
        {"$set": {"approval_status": "Rejected", "is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"success": True, "message": "Customer rejected"}

@app.patch("/api/customers/{customer_id}/status")
async def toggle_customer_status(customer_id: int, is_active: bool = Body(..., embed=True), payload: dict = Depends(verify_token)):
    result = db.customers.update_one(
        {"id": customer_id},
        {"$set": {"is_active": is_active, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"success": True, "message": f"Customer {'activated' if is_active else 'deactivated'}"}

@app.delete("/api/customers/{customer_id}")
async def delete_customer(customer_id: int, payload: dict = Depends(verify_token)):
    result = db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"success": True, "message": "Customer deleted"}

@app.post("/api/customers/{customer_id}/archive")
async def archive_customer(customer_id: int, payload: dict = Depends(verify_token)):
    result = db.customers.update_one(
        {"id": customer_id},
        {"$set": {"approval_status": "Archived", "is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"success": True, "message": "Customer archived"}

# ============ OLD PRODUCT ROUTES (for admin panel compatibility) ============

@app.get("/api/products")
async def get_products(
    search: Optional[str] = None,
    category: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"id": {"$regex": search, "$options": "i"}}
        ]
    if category:
        query["category"] = category
    
    products = list(db.products.find(query, {"_id": 0}))
    return {"products": products, "data": products}

@app.get("/api/products/{product_id}")
async def get_product(product_id: str, payload: dict = Depends(verify_token)):
    product = db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"product": product}

@app.post("/api/products")
async def create_product(product: dict = Body(...), payload: dict = Depends(verify_token)):
    if not product.get("id"):
        prefix = "PLY" if product.get("category") == "Plywood" else "TIM"
        count = db.products.count_documents({"category": product.get("category")})
        product["id"] = f"{prefix}-{count + 1:03d}"
    
    existing = db.products.find_one({"id": product["id"]})
    if existing:
        raise HTTPException(status_code=400, detail="Product ID already exists")
    
    db.products.insert_one(product)
    product.pop("_id", None)
    return {"success": True, "product": product}

@app.put("/api/products/{product_id}")
async def update_product(product_id: str, product: dict = Body(...), payload: dict = Depends(verify_token)):
    existing = db.products.find_one({"id": product_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    
    product.pop("_id", None)
    product.pop("id", None)
    
    db.products.update_one({"id": product_id}, {"$set": product})
    updated = db.products.find_one({"id": product_id}, {"_id": 0})
    return {"success": True, "product": updated}

@app.delete("/api/products/{product_id}")
async def delete_product(product_id: str, payload: dict = Depends(verify_token)):
    result = db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"success": True, "message": "Product deleted"}

# ============ OLD ORDER ROUTES (for admin panel compatibility) ============

@app.get("/api/orders")
async def get_orders(
    page: int = 1,
    per_page: int = 10,
    search: Optional[str] = None,
    status: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    query = {}
    
    user_id = payload.get("user_id")
    if user_id.startswith("customer_"):
        customer_id = int(user_id.replace("customer_", ""))
        query["customer_id"] = customer_id
    
    if search:
        query["$or"] = [
            {"id": {"$regex": search, "$options": "i"}},
            {"customerName": {"$regex": search, "$options": "i"}}
        ]
    
    if status and status != "all":
        query["status"] = status
    
    total = db.orders.count_documents(query)
    skip = (page - 1) * per_page
    
    orders = list(db.orders.find(query, {"_id": 0}).sort("order_date", -1).skip(skip).limit(per_page))
    
    return {
        "data": orders,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }

@app.get("/api/orders/{order_id}")
async def get_order(order_id: str, payload: dict = Depends(verify_token)):
    order = db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order

@app.post("/api/orders/{order_id}/status")
async def update_order_status(order_id: str, status: str = Body(..., embed=True), payload: dict = Depends(verify_token)):
    result = db.orders.update_one({"id": order_id}, {"$set": {"status": status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"success": True, "message": f"Order status updated to {status}"}

# ============ OLD INVOICE ROUTES ============

@app.get("/api/invoices")
async def get_invoices(
    page: int = 1,
    per_page: int = 10,
    search: Optional[str] = None,
    status: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    query = {}
    
    user_id = payload.get("user_id")
    if user_id.startswith("customer_"):
        customer_id = int(user_id.replace("customer_", ""))
        query["customer_id"] = customer_id
    
    if search:
        query["$or"] = [
            {"id": {"$regex": search, "$options": "i"}},
            {"order_id": {"$regex": search, "$options": "i"}},
            {"customerName": {"$regex": search, "$options": "i"}}
        ]
    
    if status and status != "all":
        query["status"] = status
    
    total = db.invoices.count_documents(query)
    skip = (page - 1) * per_page
    
    invoices = list(db.invoices.find(query, {"_id": 0}).sort("issue_date", -1).skip(skip).limit(per_page))
    
    return {
        "data": invoices,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }

@app.get("/api/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, payload: dict = Depends(verify_token)):
    invoice = db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return invoice

@app.put("/api/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, status: str = Body(..., embed=True), payload: dict = Depends(verify_token)):
    result = db.invoices.update_one({"id": invoice_id}, {"$set": {"status": status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"success": True, "message": f"Invoice status updated to {status}"}

# ============ ADMIN ROUTES ============

@app.get("/api/admin")
async def admin_resource(
    resource: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
    status: Optional[str] = None,
    search: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if resource == "dashboard":
        pending_orders = db.orders.count_documents({"status": {"$in": ["Created", "Pending"]}})
        week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
        new_orders_week = db.orders.count_documents({"order_date": {"$gte": week_ago}})
        due_invoices = db.invoices.count_documents({"status": "Pending"})
        pending_customers = db.customers.count_documents({"approval_status": "Pending"})
        total_customers = db.customers.count_documents({})
        active_customers = db.customers.count_documents({"approval_status": "Approved", "is_active": True})
        
        return {
            "pending_orders_count": pending_orders,
            "new_orders_week": new_orders_week,
            "due_invoices_count": due_invoices,
            "pending_customers": pending_customers,
            "total_customers": total_customers,
            "active_customers": active_customers
        }
    
    if resource == "orders":
        # Return orders from orders_v2 collection (new order system)
        query = {}
        if status and status != "all" and status != "All":
            query["status"] = status
        if search:
            query["$or"] = [
                {"id": {"$regex": search, "$options": "i"}},
                {"customerName": {"$regex": search, "$options": "i"}}
            ]
        
        total = db.orders_v2.count_documents(query)
        skip = (page - 1) * per_page
        
        # Sort pending first, then by created_at desc
        pipeline = [
            {"$match": query},
            {"$addFields": {"sort_priority": {"$cond": [{"$eq": ["$status", "Pending"]}, 0, 1]}}},
            {"$sort": {"sort_priority": 1, "created_at": -1}},
            {"$skip": skip},
            {"$limit": per_page},
            {"$project": {"_id": 0, "sort_priority": 0}}
        ]
        orders = list(db.orders_v2.aggregate(pipeline))
        
        return {
            "data": orders,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": (total + per_page - 1) // per_page
            }
        }
    
    if resource == "customers":
        query = {}
        if status and status != "all":
            if status == "active":
                query["approval_status"] = "Approved"
                query["is_active"] = True
            elif status == "inactive":
                query["approval_status"] = "Approved"
                query["is_active"] = False
            elif status == "pending":
                query["approval_status"] = "Pending"
            elif status == "archived":
                query["approval_status"] = "Archived"
        
        if search:
            query["$or"] = [
                {"name": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}},
                {"phone": {"$regex": search, "$options": "i"}},
                {"contactPerson": {"$regex": search, "$options": "i"}}
            ]
        
        total = db.customers.count_documents(query)
        skip = (page - 1) * per_page
        
        # Sort pending first
        pipeline = [
            {"$match": query},
            {"$addFields": {"sort_priority": {"$cond": [{"$eq": ["$approval_status", "Pending"]}, 0, 1]}}},
            {"$sort": {"sort_priority": 1, "created_at": -1}},
            {"$skip": skip},
            {"$limit": per_page},
            {"$project": {"_id": 0, "password": 0, "mpin": 0, "sort_priority": 0}}
        ]
        customers = list(db.customers.aggregate(pipeline))
        
        return {
            "data": customers,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": (total + per_page - 1) // per_page
            }
        }
    
    if resource == "invoices":
        query = {}
        if status and status != "all":
            query["status"] = status
        
        total = db.invoices_v2.count_documents(query)
        skip = (page - 1) * per_page
        
        invoices = list(db.invoices_v2.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page))
        
        return {
            "data": invoices,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": (total + per_page - 1) // per_page
            }
        }
    
    raise HTTPException(status_code=400, detail="Invalid resource")

@app.post("/api/admin")
async def admin_action(
    action: Optional[str] = None,
    data: dict = Body(None),
    payload: dict = Depends(verify_token)
):
    role = payload.get("role", "").lower()
    if role not in ["super admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if action == "update_order_status":
        order_id = data.get("order_id")
        status = data.get("status")
        result = db.orders.update_one({"id": order_id}, {"$set": {"status": status}})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Order not found")
        return {"success": True}
    
    raise HTTPException(status_code=400, detail="Invalid action")

# ============ USER MANAGEMENT (Super Admin Only) ============

class UserCreate(BaseModel):
    name: str = Field(..., min_length=1)
    email: str = Field(..., min_length=1)
    password: str = Field(..., min_length=6)
    phone: Optional[str] = None
    role: str = Field(..., pattern="^(Super Admin|Admin|Manager|Sales Person)$")

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None

def require_super_admin(payload: dict):
    role = payload.get("role", "").lower()
    if role != "super admin":
        raise HTTPException(status_code=403, detail="Super Admin access required")
    return payload

@app.get("/api/users")
async def list_users(
    page: int = 1,
    per_page: int = 10,
    search: Optional[str] = None,
    role: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    require_super_admin(payload)
    
    skip = (page - 1) * per_page
    query = {"role": {"$in": ["Super Admin", "Admin", "Manager", "Sales Person"]}}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]
    
    if role and role != "all":
        query["role"] = role
    
    total = db.users.count_documents(query)
    users = list(db.users.find(query, {"_id": 0, "password": 0}).sort("created_at", -1).skip(skip).limit(per_page))
    
    for user in users:
        if "_id" in user:
            user["id"] = str(user.pop("_id"))
    
    return {
        "data": users,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }

@app.post("/api/users")
async def create_user(data: UserCreate, payload: dict = Depends(verify_token)):
    require_super_admin(payload)
    
    existing = db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    new_user = {
        "email": data.email,
        "password": hash_password(data.password),
        "name": data.name,
        "role": data.role,
        "phone": data.phone or "",
        "approval_status": "Approved",
        "is_active": True,
        "pricing_tier": 1,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = db.users.insert_one(new_user)
    new_user["id"] = str(result.inserted_id)
    del new_user["password"]
    del new_user["_id"]
    
    return {"success": True, "message": "User created successfully", "user": new_user}

@app.put("/api/users/{user_email}")
async def update_user(user_email: str, data: UserUpdate, payload: dict = Depends(verify_token)):
    require_super_admin(payload)
    
    existing = db.users.find_one({"email": user_email})
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user_email == payload.get("user_id") and data.role and data.role != existing.get("role"):
        raise HTTPException(status_code=400, detail="Cannot change your own role")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if data.name is not None:
        update_data["name"] = data.name
    if data.email is not None and data.email != user_email:
        if db.users.find_one({"email": data.email}):
            raise HTTPException(status_code=400, detail="Email already in use")
        update_data["email"] = data.email
    if data.phone is not None:
        update_data["phone"] = data.phone
    if data.role is not None:
        update_data["role"] = data.role
    if data.is_active is not None:
        update_data["is_active"] = data.is_active
    
    db.users.update_one({"email": user_email}, {"$set": update_data})
    
    updated_user = db.users.find_one({"email": data.email or user_email}, {"password": 0})
    updated_user["id"] = str(updated_user.pop("_id"))
    
    return {"success": True, "message": "User updated successfully", "user": updated_user}

@app.delete("/api/users/{user_email}")
async def delete_user(user_email: str, payload: dict = Depends(verify_token)):
    require_super_admin(payload)
    
    existing = db.users.find_one({"email": user_email})
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user_email == payload.get("user_id"):
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    if existing.get("role") == "Super Admin":
        super_admin_count = db.users.count_documents({"role": "Super Admin"})
        if super_admin_count <= 1:
            raise HTTPException(status_code=400, detail="Cannot delete the last Super Admin")
    
    db.users.delete_one({"email": user_email})
    
    return {"success": True, "message": "User deleted successfully"}

# ============ SALES PORTAL ROUTES ============

@app.get("/api/sales/customers")
async def get_sales_customers(
    page: int = 1,
    per_page: int = 10,
    search: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    """Get all approved customers for sales to place orders"""
    query = {"approval_status": "Approved", "is_active": True}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"contactPerson": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]
    
    total = db.customers.count_documents(query)
    skip = (page - 1) * per_page
    
    customers = list(db.customers.find(query, {"_id": 0, "password": 0}).sort("name", 1).skip(skip).limit(per_page))
    
    return {
        "data": customers,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }

@app.post("/api/sales/customers")
async def create_sales_customer(customer: AdminCustomerCreate, payload: dict = Depends(verify_token)):
    """Sales person creates a customer - auto approved"""
    role = payload.get("role", "").lower()
    if role not in ["sales person", "super admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized to create customers")
    
    existing = db.customers.find_one({"email": customer.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    existing_phone = db.customers.find_one({"phone": customer.phone})
    if existing_phone:
        raise HTTPException(status_code=400, detail="Phone number already exists")
    
    # Get sales person info
    sales_person = db.users.find_one({"email": payload.get("user_id")})
    
    new_id = get_next_customer_id()
    new_customer = {
        "id": new_id,
        "email": customer.email,
        "name": customer.name,
        "business_name": customer.name,
        "contactPerson": customer.contactPerson,
        "phone": customer.phone,
        "role": "Customer",
        "approval_status": "Approved",  # Auto approved
        "is_active": True,
        "pricing_tier": customer.pricing_tier,
        "outstanding_balance": 0,
        "credit_limit": customer.credit_limit or 25000,
        "gst_number": customer.gst_number,
        "address": customer.address,
        "city": customer.city,
        "state": customer.state,
        "pincode": customer.pincode,
        "notes": customer.notes,
        "sales_person_id": str(sales_person.get("_id", "")) if sales_person else None,
        "sales_person_name": sales_person.get("name") if sales_person else None,
        "password": hash_password("customer123"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    db.customers.insert_one(new_customer)
    
    # Return without sensitive data
    del new_customer["_id"]
    del new_customer["password"]
    
    return {"success": True, "message": "Customer created successfully", "customer": new_customer}

@app.get("/api/sales/dashboard")
async def get_sales_dashboard(payload: dict = Depends(verify_token)):
    """Dashboard for sales person"""
    user_id = payload.get("user_id")
    
    # Count orders placed by this sales person
    total_orders = db.orders_v2.count_documents({"placed_by_role": "Sales Person"})
    pending_orders = db.orders_v2.count_documents({"placed_by_role": "Sales Person", "status": "Pending"})
    
    # Get recent orders
    recent_orders = list(db.orders_v2.find({"placed_by_role": "Sales Person"}, {"_id": 0}).sort("created_at", -1).limit(5))
    
    return {
        "total_orders": total_orders,
        "pending_orders": pending_orders,
        "recent_orders": recent_orders
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
